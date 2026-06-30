# -*- coding: utf-8 -*-
"""공식 -> 결과값 도출 과정을 엑셀로 정리 + 개념 다이어그램 생성.

실제 rail_analyzer 모듈을 그대로 호출해 '진짜 숫자'로 단계별 예제를 채운다.
산출물:
    설명자료/레일하중_공식설명.xlsx
    설명자료/img/*.png  (개념도)
"""
from __future__ import annotations

import os
import numpy as np

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import FancyArrowPatch, FancyBboxPatch, Arc

from rail_analyzer.geometry import build_geometry
from rail_analyzer.dynamics import (Pendulum, simulate, tension, node_force,
                                    _g_eff, _Z, GRAVITY)
from rail_analyzer.loads import MovingDevice, traction_accel, device_force_at
from rail_analyzer.driving import run_case, _centripetal, _steady_dir

# ----------------------------------------------------------------------------
# 한글 폰트
# ----------------------------------------------------------------------------
for fam in ("Malgun Gothic", "맑은 고딕", "NanumGothic"):
    try:
        plt.rcParams["font.family"] = fam
        break
    except Exception:
        pass
plt.rcParams["axes.unicode_minus"] = False

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(HERE, "설명자료")
IMG_DIR = os.path.join(OUT_DIR, "img")
os.makedirs(IMG_DIR, exist_ok=True)

# ============================================================================
# 1. 입력값 (GUI 기본값)
# ============================================================================
M_PERSON = 150.0      # kg
M_TROLLEY = 30.0      # kg
L = 1.5               # m  줄 길이
V_MAX = 2.0           # m/s
POWER_W = 10_000.0    # W (10 kW)
BASE_SPEED = 0.6      # m/s
EFF = 0.85
BRAKE = 4.0           # m/s^2
ACCEL_LIMIT = 2.0     # m/s^2
DAMPING = 0.03
SEG = 0.3             # m

g = GRAVITY

# ============================================================================
# 2. 샘플 경로: 직선 5m -> 반경 3m 90도 곡선 -> 직선 3m  (평면 XY)
# ============================================================================
def make_path():
    pts = []
    # 직선 (X 방향)
    for x in np.linspace(0, 5, 6):
        pts.append([x, 0, 0])
    # 90도 곡선, 중심 (5,3), 반경 3
    cx, cy, R = 5.0, 3.0, 3.0
    for ang in np.linspace(-np.pi / 2, 0, 12)[1:]:
        pts.append([cx + R * np.cos(ang), cy + R * np.sin(ang), 0])
    # 곡선 끝에서 +Y 직선
    for y in np.linspace(3, 6, 5)[1:]:
        pts.append([8, y, 0])
    return np.array(pts, dtype=float)


geom = build_geometry(make_path(), SEG)
pen = Pendulum(M_PERSON, M_TROLLEY, L, DAMPING)
device = MovingDevice(mass=M_PERSON + M_TROLLEY, speed=V_MAX,
                      power=POWER_W, base_speed=BASE_SPEED, efficiency=EFF)

f_max, a_raw = traction_accel(M_PERSON + M_TROLLEY, POWER_W, BASE_SPEED, EFF)
a_acc = min(a_raw, ACCEL_LIMIT)

# 케이스 실행 (전체 노드)
res_acc = run_case(geom, pen, "acc", V_MAX, a_acc, BRAKE)
res_con = run_case(geom, pen, "con", V_MAX, a_acc, BRAKE)
res_dec = run_case(geom, pen, "dec", V_MAX, a_acc, BRAKE)

# 대표 노드 선택
i_straight = 8                                   # 직선 가속 구간
i_curve = int(np.argmin(geom.radius))            # 곡률 최소(가장 휜) 노드


def node_dump(i):
    return dict(
        coord=geom.nodes[i], s=float(geom.s[i]),
        t=geom.tangent[i], n=geom.normal[i], R=float(geom.radius[i]),
    )


# ============================================================================
# 3. 단계별 손계산 (정상상태 닫힌식) - 곡선 등속(CON) 노드
# ============================================================================
def hand_calc_con(i):
    """곡선 등속: 정상상태 닫힌식으로 단계별 값 산출."""
    R = geom.radius[i]
    tvec, nvec = geom.tangent[i], geom.normal[i]
    v = V_MAX
    a_c = (v * v / R) * nvec                      # 원심 가속도 벡터
    a_trolley = a_c.copy()
    g_eff = -g * _Z - a_trolley                   # 유효중력
    n_dir = g_eff / np.linalg.norm(g_eff)         # 정상상태 줄방향
    T = M_PERSON * np.dot(g_eff, n_dir)           # 정상상태 장력(ndot=0)
    F = -M_TROLLEY * a_trolley - M_TROLLEY * g * _Z + T * n_dir
    theta = np.degrees(np.arccos(np.clip(np.dot(-n_dir, _Z), -1, 1)))
    return dict(R=R, v=v, a_c=a_c, a_trolley=a_trolley, g_eff=g_eff,
                n_dir=n_dir, T=T, F=F, theta=theta,
                ac_mag=np.linalg.norm(a_c))


hc = hand_calc_con(i_curve)

# 동역학 시뮬 결과(코드가 실제로 쓰는 값, 과도 포함 최대)
sim_con_F = res_con.node_force[i_curve]
sim_con_T = res_con.tension[i_curve]
sim_con_daf = res_con.daf[i_curve]

# 정역학(loads.py) 비교
device.accel = 0.0  # 등속
stat_con_F = device_force_at(device, geom, i_curve)

print("a_acc =", a_acc, "f_max =", f_max)
print("curve node", i_curve, "R=", hc["R"])
print("hand F =", hc["F"], "  sim F =", sim_con_F)

# ============================================================================
# 4. 개념 다이어그램
# ============================================================================
ARROW = dict(arrowstyle="-|>", mutation_scale=18, lw=2.2)


def fig_coordinate():
    """좌표계 + 곡선 레일 위 접선/법선/중력."""
    fig, ax = plt.subplots(figsize=(7.2, 5.2))
    # 곡선 레일 (평면도 느낌)
    th = np.linspace(-0.5, 1.3, 100)
    rx, ry = 5 + 3 * np.cos(th - np.pi / 2), 3 + 3 * np.sin(th - np.pi / 2)
    ax.plot(rx, ry, color="#444", lw=3, label="레일 경로")
    # 한 점
    k = 55
    px, py = rx[k], ry[k]
    tang = np.array([rx[k + 1] - rx[k - 1], ry[k + 1] - ry[k - 1]])
    tang = tang / np.linalg.norm(tang)
    norm = np.array([5 - px, 3 - py]); norm = norm / np.linalg.norm(norm)
    ax.plot(px, py, "o", color="#d62728", ms=10, zorder=5)
    ax.add_patch(FancyArrowPatch((px, py), (px + 1.6 * tang[0], py + 1.6 * tang[1]),
                                 color="#1f77b4", **ARROW))
    ax.text(px + 1.7 * tang[0], py + 1.7 * tang[1], "접선 t\n(진행방향)",
            color="#1f77b4", fontsize=11, ha="left", va="center")
    ax.add_patch(FancyArrowPatch((px, py), (px + 1.6 * norm[0], py + 1.6 * norm[1]),
                                 color="#2ca02c", **ARROW))
    ax.text(px + 1.7 * norm[0], py + 1.7 * norm[1], "법선 n\n(곡률중심)",
            color="#2ca02c", fontsize=11, ha="right", va="center")
    # 곡률중심
    ax.plot(5, 3, "x", color="#2ca02c", ms=10)
    ax.text(5.1, 3.1, "곡률중심 (반경 R)", color="#2ca02c", fontsize=9)
    # 전역 좌표축
    ox, oy = 0.3, 5.5
    ax.add_patch(FancyArrowPatch((ox, oy), (ox + 1.2, oy), color="#666", **ARROW))
    ax.add_patch(FancyArrowPatch((ox, oy), (ox, oy + 1.0), color="#666", **ARROW))
    ax.text(ox + 1.3, oy, "X", fontsize=11); ax.text(ox, oy + 1.1, "Y", fontsize=11)
    ax.text(ox, oy - 0.45, "Z: 화면 밖(연직 위)", fontsize=9, color="#666")
    ax.set_title("① 경로 기하 — 모든 힘의 '방향'은 여기서 나온다", fontsize=13, weight="bold")
    ax.set_aspect("equal"); ax.axis("off")
    ax.legend(loc="lower right", fontsize=9)
    p = os.path.join(IMG_DIR, "1_geometry.png")
    fig.tight_layout(); fig.savefig(p, dpi=130); plt.close(fig)
    return p


def fig_pendulum():
    """진자 자유물체도: 트롤리 가속 -> 줄 기울기 -> 장력 분해."""
    fig, ax = plt.subplots(figsize=(7.2, 6.0))
    # 레일
    ax.plot([-1, 4], [4, 4], color="#444", lw=4)
    ax.text(3.6, 4.15, "레일", fontsize=10)
    # 트롤리
    tx, ty = 1.5, 4.0
    ax.add_patch(FancyBboxPatch((tx - 0.35, ty - 0.18), 0.7, 0.36,
                 boxstyle="round,pad=0.02", fc="#9ecae1", ec="#1f77b4", lw=1.5))
    ax.text(tx, ty + 0.4, "트롤리 (m_t)", ha="center", fontsize=10, color="#1f77b4")
    # 기울어진 줄
    ang = np.radians(25)  # 연직에서 기운 각
    Lp = 2.6
    bx, by = tx + Lp * np.sin(ang), ty - Lp * np.cos(ang)
    ax.plot([tx, bx], [ty, by], color="#8c564b", lw=2.5)
    # 사람(질점)
    ax.plot(bx, by, "o", color="#d62728", ms=22, zorder=5)
    ax.text(bx + 0.35, by, "사람 (m_p)", fontsize=10, color="#d62728")
    # 연직 점선
    ax.plot([tx, tx], [ty, ty - Lp], "--", color="#aaa", lw=1.2)
    ax.add_patch(Arc((tx, ty), 1.2, 1.2, angle=0, theta1=270, theta2=270 + 25,
                     color="#555"))
    ax.text(tx + 0.35, ty - 0.85, "θ", fontsize=13, color="#555")
    # 트롤리 가속도 화살표
    ax.add_patch(FancyArrowPatch((tx, ty + 0.0), (tx + 1.3, ty), color="#ff7f0e",
                 **ARROW))
    ax.text(tx + 1.35, ty - 0.05, "a_trolley (가속/원심)", color="#ff7f0e",
            fontsize=10, va="center")
    # 장력(줄방향, 사람->트롤리)
    tdir = np.array([tx - bx, ty - by]); tdir = tdir / np.linalg.norm(tdir)
    ax.add_patch(FancyArrowPatch((bx, by), (bx + 1.5 * tdir[0], by + 1.5 * tdir[1]),
                 color="#2ca02c", **ARROW))
    ax.text(bx + 1.5 * tdir[0] + 0.05, by + 1.5 * tdir[1], "장력 T", color="#2ca02c",
            fontsize=11)
    # 중력
    ax.add_patch(FancyArrowPatch((bx, by), (bx, by - 1.3), color="#d62728", **ARROW))
    ax.text(bx + 0.1, by - 1.3, "m_p·g", color="#d62728", fontsize=10)
    # 유효중력
    geff = np.array([-np.sin(ang), -np.cos(ang)])
    ax.add_patch(FancyArrowPatch((bx, by), (bx + 1.3 * geff[0], by + 1.3 * geff[1]),
                 color="#9467bd", **ARROW, ls="--"))
    ax.text(bx + 1.3 * geff[0] - 0.1, by + 1.3 * geff[1] - 0.2,
            "유효중력 g_eff\n=-g·Z - a_trolley", color="#9467bd", fontsize=9, ha="right")
    ax.set_title("② 매달린 사람 = 구면진자 — 가속하면 줄이 기울고 장력이 커진다",
                 fontsize=12.5, weight="bold")
    ax.set_xlim(-1.2, 5.5); ax.set_ylim(0.3, 5.2)
    ax.set_aspect("equal"); ax.axis("off")
    p = os.path.join(IMG_DIR, "2_pendulum.png")
    fig.tight_layout(); fig.savefig(p, dpi=130); plt.close(fig)
    return p


def fig_flow():
    """도출 플로우 다이어그램."""
    fig, ax = plt.subplots(figsize=(9.5, 6.2))
    ax.axis("off")
    ax.set_xlim(0, 10); ax.set_ylim(0, 10)

    def box(x, y, w, h, text, fc, ec):
        ax.add_patch(FancyBboxPatch((x, y), w, h, boxstyle="round,pad=0.08",
                     fc=fc, ec=ec, lw=1.8))
        ax.text(x + w / 2, y + h / 2, text, ha="center", va="center",
                fontsize=9.5)

    def arrow(x1, y1, x2, y2):
        ax.add_patch(FancyArrowPatch((x1, y1), (x2, y2), color="#555",
                     arrowstyle="-|>", mutation_scale=16, lw=1.8))

    box(0.3, 7.6, 3.2, 1.6,
        "입력\n사람·트롤리 질량, 줄길이 L,\n최고속도 v, 출력 P, 감속 a_b", "#fff2cc", "#d6b656")
    box(0.3, 4.8, 3.2, 1.6,
        "기하 (DXF 경로)\n노드별 접선 t, 법선 n,\n곡률반경 R", "#d5e8d4", "#82b366")
    arrow(1.9, 7.6, 1.9, 6.4)

    box(4.4, 7.6, 2.6, 1.6,
        "주행 가속도\na = a_tan·t + (v²/R)·n", "#dae8fc", "#6c8ebf")
    arrow(3.5, 8.4, 4.4, 8.4)

    box(4.4, 4.8, 2.6, 1.6,
        "진자 동역학\ng_eff = -g·Z - a\n줄방향 n, 장력 T (RK4)", "#dae8fc", "#6c8ebf")
    arrow(5.7, 7.6, 5.7, 6.4)
    arrow(3.5, 5.6, 4.4, 5.6)

    box(7.7, 6.0, 2.1, 2.0,
        "노드 힘 F\n= -m_t·a\n  - m_t·g·Z\n  + T·n", "#f8cecc", "#b85450")
    arrow(7.0, 5.6, 7.7, 6.6)
    arrow(7.0, 8.4, 7.7, 7.6)

    box(7.4, 2.8, 2.7, 1.7,
        "결과: 노드별 [Fx,Fy,Fz]\nACC / CON / DEC 중 최대\n→ MGT 하중", "#e1d5e7", "#9673a6")
    arrow(8.75, 6.0, 8.75, 4.5)

    box(0.3, 2.0, 3.2, 1.6,
        "정역학(등가정적)\nW=m·g·DAF (↓Z)\nF_t=m·a (t)\nF_c=m·v²/R (n)", "#d5e8d4", "#82b366")
    arrow(1.9, 4.8, 1.9, 3.6)
    ax.text(1.9, 1.7, "※ 진자 대신 1점질량으로 단순합 (비교용)",
            ha="center", fontsize=8, color="#666")

    ax.set_title("③ X·Y·Z 힘이 도출되는 전체 흐름", fontsize=13, weight="bold")
    p = os.path.join(IMG_DIR, "3_flow.png")
    fig.tight_layout(); fig.savefig(p, dpi=130); plt.close(fig)
    return p


img1 = fig_coordinate()
img2 = fig_pendulum()
img3 = fig_flow()
print("images done")

# ============================================================================
# 5. 엑셀 작성
# ============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

wb = Workbook()

# 스타일
TITLE = Font(name="맑은 고딕", size=15, bold=True, color="1F3864")
H = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
SUB = Font(name="맑은 고딕", size=11, bold=True, color="1F3864")
BODY = Font(name="맑은 고딕", size=10)
MONO = Font(name="Consolas", size=10)
NUMB = Font(name="Consolas", size=10, bold=True, color="C00000")
HFILL = PatternFill("solid", fgColor="2F5496")
SUBFILL = PatternFill("solid", fgColor="D9E1F2")
YFILL = PatternFill("solid", fgColor="FFF2CC")
GFILL = PatternFill("solid", fgColor="E2EFDA")
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="center")
CTR = Alignment(horizontal="center", vertical="center")


def vec(v):
    return f"({v[0]:.3f}, {v[1]:.3f}, {v[2]:.3f})"


def style_header_row(ws, row, ncol, fill=HFILL, font=H):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = font
        cell.alignment = CTR; cell.border = BORD


# ---------------------------------------------------------------- 시트 0: 개념도
ws = wb.active
ws.title = "0_개념도"
ws.sheet_view.showGridLines = False
ws["A1"] = "레일 이동하중 — 공식에 의한 X·Y·Z 힘 도출 (개념)"
ws["A1"].font = TITLE
ws["A2"] = "매달린 사람을 '구면진자'로 보고, 트롤리 주행 가속이 진자를 흔들 때 레일 부착점에 걸리는 3방향 힘을 구한다."
ws["A2"].font = BODY
row = 4
for img, hcell in [(img1, "A"), (img2, "A"), (img3, "A")]:
    im = XLImage(img)
    im.anchor = f"{hcell}{row}"
    ws.add_image(im)
    row += 30
ws.column_dimensions["A"].width = 14

# ---------------------------------------------------------------- 시트 1: 입력값
ws = wb.create_sheet("1_입력값")
ws.sheet_view.showGridLines = False
ws["A1"] = "1. 입력값 (GUI 기본값)"; ws["A1"].font = TITLE
hdr = ["기호", "항목", "값", "단위", "비고"]
ws.append([])
ws.append(hdr); style_header_row(ws, 3, 5)
rows = [
    ("m_p", "사람(운전자) 질량", M_PERSON, "kg", "하네스 포함"),
    ("m_t", "트롤리(동력체) 질량", M_TROLLEY, "kg", ""),
    ("L", "줄 길이", L, "m", "진자 길이"),
    ("v_max", "최고속도", V_MAX, "m/s", "원심력에 사용"),
    ("P", "추진출력", POWER_W, "W", "= 10 kW"),
    ("v_b", "기저속도", BASE_SPEED, "m/s", "일정견인력 한계"),
    ("η", "구동효율", EFF, "-", ""),
    ("a_b", "비상감속", BRAKE, "m/s²", "DEC 케이스"),
    ("a_lim", "가속 상한", ACCEL_LIMIT, "m/s²", "안락도 제한 ≈0.2g"),
    ("ζ", "진자 감쇠비", DAMPING, "-", "공기저항 등"),
    ("g", "중력가속도", g, "m/s²", "상수"),
]
for r in rows:
    ws.append(list(r))
for rr in range(4, 4 + len(rows)):
    for cc in range(1, 6):
        ws.cell(rr, cc).border = BORD
        ws.cell(rr, cc).font = BODY
    ws.cell(rr, 3).font = NUMB
    ws.cell(rr, 1).font = MONO
for w, c in zip([8, 22, 12, 8, 22], "ABCDE"):
    ws.column_dimensions[c].width = w

# ---------------------------------------------------------------- 시트 2: 견인가속
ws = wb.create_sheet("2_견인가속도")
ws.sheet_view.showGridLines = False
ws["A1"] = "2. 마력으로부터 최대 가속도 산정"; ws["A1"].font = TITLE
ws["A3"] = "공식"; ws["A3"].font = SUB; ws["A3"].fill = SUBFILL
ws["A4"] = "최대견인력  F_max = η · P / v_b"; ws["A4"].font = MONO
ws["A5"] = "견인가속도  a_raw = F_max / (m_p + m_t)"; ws["A5"].font = MONO
ws["A6"] = "적용가속도  a_acc = min(a_raw, a_lim)   ← 사람운반 안락도 제한"; ws["A6"].font = MONO
ws["A8"] = "대입"; ws["A8"].font = SUB; ws["A8"].fill = SUBFILL
calc = [
    ("F_max", f"{EFF} × {POWER_W:.0f} / {BASE_SPEED}", f_max, "N"),
    ("a_raw", f"{f_max:.1f} / ({M_PERSON}+{M_TROLLEY})", a_raw, "m/s²"),
    ("a_acc", f"min({a_raw:.3f}, {ACCEL_LIMIT})", a_acc, "m/s²"),
]
ws.append([]) if False else None
r0 = 9
ws.cell(r0, 1, "기호").font = H; ws.cell(r0, 2, "대입식").font = H
ws.cell(r0, 3, "결과").font = H; ws.cell(r0, 4, "단위").font = H
style_header_row(ws, r0, 4)
for k, (sym, expr, val, unit) in enumerate(calc):
    rr = r0 + 1 + k
    ws.cell(rr, 1, sym).font = MONO
    ws.cell(rr, 2, expr).font = MONO
    ws.cell(rr, 3, round(val, 4)).font = NUMB
    ws.cell(rr, 4, unit).font = BODY
    for cc in range(1, 5):
        ws.cell(rr, cc).border = BORD
ws.cell(r0 + 5, 1, "→ a_raw 가 상한을 넘으므로 a_acc = 2.0 m/s² 로 클램프됨").font = BODY
for w, c in zip([10, 34, 12, 8], "ABCD"):
    ws.column_dimensions[c].width = w

# ---------------------------------------------------------------- 시트 3: 기하
ws = wb.create_sheet("3_기하(방향)")
ws.sheet_view.showGridLines = False
ws["A1"] = "3. 경로 기하 — 힘의 '방향'을 정함"; ws["A1"].font = TITLE
ws["A2"] = "DXF 경로를 노드로 쪼개고, 노드마다 접선 t·법선 n·곡률반경 R 을 구한다."; ws["A2"].font = BODY
ws["A4"] = "공식"; ws["A4"].font = SUB; ws["A4"].fill = SUBFILL
gf = [
    "접선 t  = 인접 두 세그먼트 단위벡터의 평균 (진행방향)",
    "곡률각 θ = arccos( u_(i-1) · u_i )   (두 세그먼트 사이 꺾인 각)",
    "곡률반경 R = ds / θ        (직선부는 R = ∞)",
    "법선 n  = unit( u_i - u_(i-1) )   (곡률중심 방향, 직선부는 0)",
]
for k, t in enumerate(gf):
    ws.cell(5 + k, 1, t).font = MONO
ws.cell(10, 1, "대표 노드 값 (이 엑셀에서 쓰는 예제)").font = SUB
ws.cell(10, 1).fill = SUBFILL
hdr = ["", "노드#", "좌표(x,y,z)", "접선 t", "법선 n", "곡률반경 R[m]"]
ws.append([]) if False else None
r0 = 11
for c, h in enumerate(hdr, 1):
    ws.cell(r0, c, h)
style_header_row(ws, r0, len(hdr))
ds = node_dump(i_straight); dc = node_dump(i_curve)
ws.cell(r0 + 1, 1, "직선·가속").font = BODY
ws.cell(r0 + 1, 2, i_straight).font = MONO
ws.cell(r0 + 1, 3, vec(ds["coord"])).font = MONO
ws.cell(r0 + 1, 4, vec(ds["t"])).font = MONO
ws.cell(r0 + 1, 5, vec(ds["n"])).font = MONO
ws.cell(r0 + 1, 6, "∞ (직선)").font = NUMB
ws.cell(r0 + 2, 1, "곡선").font = BODY
ws.cell(r0 + 2, 2, i_curve).font = MONO
ws.cell(r0 + 2, 3, vec(dc["coord"])).font = MONO
ws.cell(r0 + 2, 4, vec(dc["t"])).font = MONO
ws.cell(r0 + 2, 5, vec(dc["n"])).font = MONO
ws.cell(r0 + 2, 6, round(dc["R"], 3)).font = NUMB
for rr in (r0 + 1, r0 + 2):
    for cc in range(1, len(hdr) + 1):
        ws.cell(rr, cc).border = BORD
for w, c in zip([12, 8, 20, 20, 20, 14], "ABCDEF"):
    ws.column_dimensions[c].width = w

# ---------------------------------------------------------------- 시트 4: 동역학 공식
ws = wb.create_sheet("4_동역학공식")
ws.sheet_view.showGridLines = False
ws["A1"] = "4. 동역학(구면진자) 핵심 공식"; ws["A1"].font = TITLE
blocks = [
    ("① 유효중력 (비관성 좌표계)",
     ["g_eff = -g·Z - a_trolley",
      "  트롤리가 가속하면 관성력이 더해져 '느껴지는 중력'이 비스듬해진다.",
      "  → 진자는 g_eff 방향으로 매달리려 한다."]),
    ("② 줄방향 운동방정식 (RK4로 시간적분)",
     ["n̈ = (1/L)[g_eff - (g_eff·n)n] - (ṅ·ṅ)n - 2ζωₙ·ṅ",
      "  1항: 복원(흔들림) / 2항: 줄길이 일정 구속 / 3항: 감쇠",
      "  ωₙ = √(g/L),  진자주기 = 2π√(L/g)"]),
    ("③ 줄 장력",
     ["T = m_p·[ (g_eff·n) + L·(ṅ·ṅ) ]",
      "  정상상태(흔들림 멈춤, ṅ=0): T = m_p·(g_eff·n)"]),
    ("④ 레일 노드(부착점) 힘  ★최종★",
     ["F_node = -m_t·a_trolley  -  m_t·g·Z  +  T·n",
      "  1항: 트롤리 가속 반작용(주로 수평) ",
      "  2항: 트롤리 자중(Z)",
      "  3항: 줄 장력을 줄방향 n 으로 (n이 기울어 X·Y·Z 전부 생김)"]),
]
rr = 3
for title, lines in blocks:
    ws.cell(rr, 1, title).font = SUB; ws.cell(rr, 1).fill = SUBFILL
    rr += 1
    for ln in lines:
        ws.cell(rr, 1, ln).font = MONO if rr == rr else BODY
        ws.cell(rr, 1).font = MONO if not ln.startswith("  ") else BODY
        rr += 1
    rr += 1
ws.column_dimensions["A"].width = 80

# ---------------------------------------------------------------- 시트 5: 단계별 예제
ws = wb.create_sheet("5_단계별예제")
ws.sheet_view.showGridLines = False
ws["A1"] = "5. 단계별 숫자 대입 — 곡선 등속(CON) 노드"; ws["A1"].font = TITLE
ws["A2"] = f"노드 #{i_curve}, 곡률반경 R={hc['R']:.3f} m, 속도 v={hc['v']:.1f} m/s (정상상태 닫힌식)"
ws["A2"].font = BODY
steps = [
    ("STEP 1", "원심 가속도", "a_c = v²/R · n",
     f"{hc['v']}²/{hc['R']:.3f} · n", f"|a_c|={hc['ac_mag']:.3f} m/s²  →  {vec(hc['a_c'])}"),
    ("STEP 2", "트롤리 가속도", "a_trolley = a_tan·t + a_c",
     "a_tan=0 (등속)", vec(hc['a_trolley'])),
    ("STEP 3", "유효중력", "g_eff = -g·Z - a_trolley",
     f"-{g}·Z - {vec(hc['a_trolley'])}", vec(hc['g_eff'])),
    ("STEP 4", "줄방향(정상)", "n = g_eff / |g_eff|",
     "정규화", f"{vec(hc['n_dir'])}   (연직에서 θ={hc['theta']:.2f}°)"),
    ("STEP 5", "줄 장력", "T = m_p·(g_eff·n)",
     f"{M_PERSON}·(g_eff·n)", f"T = {hc['T']:.1f} N  (정적자중 {M_PERSON*g:.0f}N의 {hc['T']/(M_PERSON*g):.3f}배)"),
    ("STEP 6", "노드 힘", "F = -m_t·a - m_t·g·Z + T·n",
     "각 항 합산", vec(hc['F'])),
    ("STEP 7", "크기", "|F| = √(Fx²+Fy²+Fz²)",
     "", f"|F| = {np.linalg.norm(hc['F']):.1f} N"),
]
r0 = 4
for c, h in enumerate(["단계", "항목", "공식", "대입", "결과값"], 1):
    ws.cell(r0, c, h)
style_header_row(ws, r0, 5)
for k, s in enumerate(steps):
    rr = r0 + 1 + k
    for cc, val in enumerate(s, 1):
        cell = ws.cell(rr, cc, val)
        cell.border = BORD
        cell.font = MONO if cc in (3, 4) else BODY
        cell.alignment = WRAP
    ws.cell(rr, 5).font = NUMB
    ws.cell(rr, 1).fill = YFILL
ws.cell(r0 + len(steps) + 2, 1,
        "검증: 코드(RK4 과도 포함 최대)와 비교").font = SUB
ws.cell(r0 + len(steps) + 2, 1).fill = SUBFILL
ws.cell(r0 + len(steps) + 3, 1, "손계산(정상상태) F").font = BODY
ws.cell(r0 + len(steps) + 3, 5, vec(hc['F'])).font = MONO
ws.cell(r0 + len(steps) + 4, 1, "코드 시뮬 F (과도 최대)").font = BODY
ws.cell(r0 + len(steps) + 4, 5, vec(sim_con_F)).font = MONO
ws.cell(r0 + len(steps) + 5, 1, "정역학(loads.py) F").font = BODY
ws.cell(r0 + len(steps) + 5, 5, vec(stat_con_F)).font = MONO
for w, c in zip([10, 14, 26, 26, 34], "ABCDE"):
    ws.column_dimensions[c].width = w

# ---------------------------------------------------------------- 시트 6: 3케이스 결과
ws = wb.create_sheet("6_3케이스결과")
ws.sheet_view.showGridLines = False
ws["A1"] = "6. 3주행케이스 노드 하중 결과"; ws["A1"].font = TITLE
ws["A2"] = "ACC(최대가속) / CON(최고등속) / DEC(비상감속) — 노드별로 셋 중 최대를 설계하중으로 채택"
ws["A2"].font = BODY


def case_block(start_row, label, res, i):
    ws.cell(start_row, 1, label).font = SUB
    ws.cell(start_row, 1).fill = SUBFILL
    hdr = ["노드#", "Fx[N]", "Fy[N]", "Fz[N]", "|F|[N]", "장력T[N]", "DAF", "속도[m/s]"]
    for c, h in enumerate(hdr, 1):
        ws.cell(start_row + 1, c, h)
    style_header_row(ws, start_row + 1, len(hdr))
    for k, idx in enumerate([i_straight, i_curve]):
        rr = start_row + 2 + k
        F = res.node_force[idx]
        vals = [idx, F[0], F[1], F[2], np.linalg.norm(F),
                res.tension[idx], res.daf[idx], res.v_node[idx]]
        for cc, v in enumerate(vals, 1):
            cell = ws.cell(rr, cc, round(float(v), 2) if cc > 1 else int(v))
            cell.border = BORD; cell.font = MONO
    return start_row + 5


r = 4
r = case_block(r, "ACC (최대가속)", res_acc, None)
r = case_block(r, "CON (최고등속)", res_con, None)
r = case_block(r, "DEC (비상감속)", res_dec, None)

# 엔벨로프
ws.cell(r, 1, "Envelope (노드별 셋 중 |F| 최대)").font = SUB
ws.cell(r, 1).fill = GFILL
hdr = ["노드#", "지배케이스", "Fx[N]", "Fy[N]", "Fz[N]", "|F|[N]"]
for c, h in enumerate(hdr, 1):
    ws.cell(r + 1, c, h)
style_header_row(ws, r + 1, len(hdr), fill=PatternFill("solid", fgColor="548235"))
for k, idx in enumerate([i_straight, i_curve]):
    rr = r + 2 + k
    forces = {"ACC": res_acc.node_force[idx], "CON": res_con.node_force[idx],
              "DEC": res_dec.node_force[idx]}
    best = max(forces, key=lambda kk: np.linalg.norm(forces[kk]))
    F = forces[best]
    vals = [idx, best, F[0], F[1], F[2], np.linalg.norm(F)]
    for cc, v in enumerate(vals, 1):
        cell = ws.cell(rr, cc, v if cc == 2 else (int(v) if cc == 1 else round(float(v), 2)))
        cell.border = BORD; cell.font = MONO
for w, c in zip([10, 12, 12, 12, 12, 12, 10, 10], "ABCDEFGH"):
    ws.column_dimensions[c].width = w

# ---------------------------------------------------------------- 저장
xlsx = os.path.join(OUT_DIR, "레일하중_공식설명.xlsx")
wb.save(xlsx)
print("SAVED:", xlsx)

